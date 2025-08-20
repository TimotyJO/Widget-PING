"""
Ping Monitor - GUI (Tkinter) + Headless Fallback
-------------------------------------------------
Update:
- Hapus kolom Priority di UI.
- Ganti label mode: **Mode: REALTIME** (bukan REAL) saat tidak dummy.
- Rekap otomatis: **Sen–Jum 15:30**, **Sabtu 13:30**, Minggu libur.
- Tambah kolom **JamDown <Target>** di Excel/CSV berisi daftar jam HH:MM saat kejadian DOWN per hari.
- Sheet **Grafik** (Excel) hanya **1 chart**: **100% Stacked Column** komposisi DOWN (30 hari). Semua chart lama dihapus.
- **Keep Awake**: saat aplikasi berjalan, sistem dicegah tidur otomatis (Windows) sehingga tetap aktif ketika layar dikunci (Win+L). Bisa dimatikan dengan env `KEEP_AWAKE=0`.
- Headless (tanpa tkinter) tetap berjalan dan rekap ke CSV bila openpyxl tidak ada.
- Unit tests tersedia (RUN_TESTS=1).
"""
from __future__ import annotations
import os
import sys
import csv
import time
import platform
import random
import subprocess
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Tuple

# ========== Optional imports & capability flags ==========
try:
    import tkinter as tk
    from tkinter import ttk, messagebox
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False
    tk = None
    ttk = None
    messagebox = None

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

# Matplotlib hanya dipakai untuk Chart GUI
try:
    import matplotlib  # type: ignore
    matplotlib.use("Agg")
    from matplotlib.figure import Figure  # type: ignore
    if TK_AVAILABLE:
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg  # type: ignore
    MPL_AVAILABLE = True
except Exception:
    MPL_AVAILABLE = False

# ======= Optional: System Tray =======
try:
    import pystray  # type: ignore
    from PIL import Image, ImageDraw  # type: ignore
    TRAY_AVAILABLE = True
except Exception:
    TRAY_AVAILABLE = False

# ======= KONFIGURASI =======
ENV_USE_DUMMY = os.environ.get("USE_DUMMY")
ENV_FORCE_REAL = os.environ.get("FORCE_REAL")
if TK_AVAILABLE:
    USE_DUMMY = (ENV_USE_DUMMY == "1")
else:
    USE_DUMMY = (ENV_FORCE_REAL != "1")  # di headless default dummy agar stabil

INTERVAL_SEC = float(os.environ.get("INTERVAL_SEC", "3.0"))
PING_COUNT = os.environ.get("PING_COUNT", "2")
POPUP_DURATION_MS = int(os.environ.get("POPUP_DURATION_MS", "2500"))
KEEP_AWAKE = os.environ.get("KEEP_AWAKE", "1") == "1"  # cegah sleep default ON di Windows

TARGETS = [
    {"name": "Server",     "host": "192.168.8.8"},
    {"name": "Desktop",    "host": "192.168.9.1"},
    {"name": "HP",         "host": "192.168.11.1"},
    {"name": "Mitraskon",  "host": "mitraskon.com"},
    {"name": "Mitralindo", "host": "mitralindo.com"},
    {"name": "Trekons",    "host": "trekons.com"},
]

REKAP_ORDER = ["Server", "Desktop", "HP", "Mitraskon", "Mitralindo", "Trekons"]

# Dummy probabilities (hanya dipakai jika USE_DUMMY=True)
P_UP = 0.78
P_WARNING = 0.15
P_DOWN = 0.07

# ======= Lokasi output (portable) =======
PCNAME = os.environ.get("COMPUTERNAME") or os.environ.get("HOSTNAME") or platform.node() or "UNKNOWN"

def _default_base_dir() -> Path:
    override = os.environ.get("PING_BASE_DIR")
    if override:
        return Path(override)
    if platform.system().lower() == "windows":
        return Path(f"C:/PING/{PCNAME}")
    return Path.home() / "PING" / PCNAME

REKAP_DIR = _default_base_dir()
REKAP_DIR.mkdir(parents=True, exist_ok=True)
REKAP_FILE_XLSX = REKAP_DIR / "rekap_ping.xlsx"
REKAP_FILE_CSV = REKAP_DIR / "rekap_ping.csv"
REKAP_FILE = REKAP_FILE_XLSX if EXCEL_AVAILABLE else REKAP_FILE_CSV

# ======= KEEP AWAKE (Windows) =======
_KEEP_AWAKE_ON = False

def set_keep_awake(enable: bool):
    """Cegah sistem tidur otomatis saat program aktif (Windows). Aman: reset saat exit."""
    global _KEEP_AWAKE_ON
    _KEEP_AWAKE_ON = enable
    if platform.system().lower() != "windows":
        return
    try:
        import ctypes
        ES_CONTINUOUS = 0x80000000
        ES_SYSTEM_REQUIRED = 0x00000001
        ES_AWAYMODE_REQUIRED = 0x00000040  # opsional, biar tetap jalan meski display dimatikan
        flags = ES_CONTINUOUS | (ES_SYSTEM_REQUIRED if enable else 0) | (ES_AWAYMODE_REQUIRED if enable else 0)
        ctypes.windll.kernel32.SetThreadExecutionState(flags)
    except Exception:
        pass

# ======= UTILITAS STATUS =======
def color_for(status: str) -> str:
    if status == "UP":
        return "#22c55e"  # green
    if status == "WARNING":
        return "#f59e0b"  # amber
    if status == "DOWN":
        return "#ef4444"  # red
    return "#9ca3af"      # gray


def dummy_status() -> str:
    r = random.random()
    if r < P_UP:
        return "UP"
    if r < P_UP + P_WARNING:
        return "WARNING"
    return "DOWN"


def check_target(host: str) -> str:
    """Ping target dan tentukan status.
    Return:
        "UP"      -> semua reply normal
        "WARNING" -> ada packet loss (sebagian)
        "DOWN"    -> unreachable/100% loss / error

    Catatan: di Windows saat dibundle menjadi .exe, panggilan ke `ping.exe`
    bisa memunculkan jendela konsol. Untuk mencegahnya kita set
    `creationflags=CREATE_NO_WINDOW` dan sembunyikan window via STARTUPINFO.
    """
    param = "-n" if platform.system().lower() == "windows" else "-c"
    try:
        kwargs = {}
        if platform.system().lower() == "windows":
            try:
                si = subprocess.STARTUPINFO()
                si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                si.wShowWindow = 0  # SW_HIDE
                kwargs["startupinfo"] = si
                kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
            except Exception:
                pass
        result = subprocess.run(
            ["ping", param, PING_COUNT, host],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=10,
            **kwargs,
        )
        out = (result.stdout or "").lower()
        if result.returncode != 0 or "unreachable" in out:
            return "DOWN"
        if ("lost = 0" in out) or ("0% packet loss" in out):
            return "UP"
        elif ("lost =" in out) or ("packet loss" in out):
            return "WARNING"
        return "DOWN"
    except Exception:
        return "DOWN"

# ======= Rekap: Excel (openpyxl) atau CSV fallback =======
def ensure_workbook(path: Path) -> tuple[Path, str]:
    """Pastikan file rekap tersedia. Return (path_file, mode): mode in {"xlsx","csv"}."""
    if EXCEL_AVAILABLE:
        try:
            if path.exists():
                wb = load_workbook(path)
                ws = wb.active
                if ws.max_row >= 1 and ws.cell(1, 1).value == "Tanggal":
                    # Jika header lama (tanpa JamDown), tambahkan kolom JamDown di ujung
                    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
                    if not any(str(h or "").startswith("JamDown ") for h in headers):
                        for name in REKAP_ORDER:
                            ws.cell(1, ws.max_column+1, f"JamDown {name}")
                        wb.save(path)
                    return path, "xlsx"
        except Exception:
            pass
        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap"
        # Header: Tanggal + Down target... + JamDown target...
        ws.append(["Tanggal"] + [f"Down {n}" for n in REKAP_ORDER] + [f"JamDown {n}" for n in REKAP_ORDER])
        wb.save(path)
        return path, "xlsx"
    else:
        if not REKAP_FILE_CSV.exists():
            with open(REKAP_FILE_CSV, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Tanggal"] + [f"Down {n}" for n in REKAP_ORDER] + [f"JamDown {n}" for n in REKAP_ORDER])
        return REKAP_FILE_CSV, "csv"

# ======= Sheet Grafik (Excel) =======
if EXCEL_AVAILABLE:
    from openpyxl.chart import LineChart, Reference

    def _get_header_map(ws) -> Dict[str, int]:
        return {str(ws.cell(1, c).value): c for c in range(1, ws.max_column+1)}

    def _parse_first_down_minutes(jam_str: str | None) -> int:
        """Ambil jam pertama (HH:MM;HH:MM;...) dan konversi ke menit. 0 jika kosong."""
        if not jam_str:
            return 0
        parts = [p.strip() for p in str(jam_str).split(";") if p.strip()]
        if not parts:
            return 0
        try:
            h, m = parts[0].split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return 0

    def update_grafik_sheet(wb):
        """Bangun sheet 'Grafik' dengan satu chart 100% Stacked Column.
        Data: komposisi Down <Target> per tanggal, ambil 30 hari terakhir.
        """
        # Pastikan sheet 'Rekap' ada
        ws_rekap = wb.active
        if ws_rekap.title != "Rekap":
            for s in wb.worksheets:
                if s.title == "Rekap":
                    ws_rekap = s
                    break
        # Buat/bersihkan sheet 'Grafik'
        if "Grafik" in wb.sheetnames:
            ws_g = wb["Grafik"]
            ws_g.delete_rows(1, ws_g.max_row or 1)
            try:
                ws_g._charts = []  # type: ignore[attr-defined]
            except Exception:
                pass
        else:
            ws_g = wb.create_sheet("Grafik")

        # Atur lebar kolom agar rapi
        for col, w in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", [12] + [10]*25):
            try:
                ws_g.column_dimensions[col].width = w
            except Exception:
                break

        # Build tabel komposisi: Tanggal + Down target
        from openpyxl.chart import BarChart, Reference
        hmap = {str(ws_rekap.cell(1, c).value): c for c in range(1, ws_rekap.max_column+1)}
        rows = list(ws_rekap.iter_rows(values_only=True))
        body = rows[1:] if len(rows) > 1 else []
        last30 = body[-30:]

        ws_g.append(["Tanggal"] + REKAP_ORDER)
        for r in last30:
            tanggal = r[0]
            vals = []
            for name in REKAP_ORDER:
                cidx = hmap.get(f"Down {name}")
                v = 0
                if cidx and cidx-1 < len(r) and isinstance(r[cidx-1], (int, float)):
                    v = int(r[cidx-1])
                vals.append(v)
            ws_g.append([tanggal] + vals)

        # Chart: 100% Stacked Column
        if ws_g.max_row >= 2:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "percentStacked"
            chart.overlap = 100
            chart.title = "Komposisi DOWN per Hari (100%) – 30 hari"
            chart.y_axis.title = "Persentase"
            chart.y_axis.number_format = "0%"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1
            chart.x_axis.title = "Tanggal"
            data_ref = Reference(ws_g, min_col=2, min_row=1, max_col=1+len(REKAP_ORDER), max_row=ws_g.max_row)
            cats_ref = Reference(ws_g, min_col=1, min_row=2, max_row=ws_g.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 20
            chart.width = 36
            ws_g.add_chart(chart, "B2")

# ======= Tulis Rekap =======
def write_rekap(today: date, counters: Dict[str, int], jam_map: Dict[str, List[str]] | None = None) -> None:
    """Tulis baris rekap. Jika Excel tersedia dan ada nilai DOWN > 0,
    buat/update sheet 'Grafik' otomatis (30 baris terakhir).
    jam_map: dict nama target -> list jam HH:MM saat DOWN hari ini.
    """
    _, mode = ensure_workbook(REKAP_FILE)
    tanggal_str = today.strftime("%d/%m/%Y")
    jam_map = jam_map or {name: [] for name in REKAP_ORDER}
    jam_joined = [";".join(jam_map.get(name, [])) for name in REKAP_ORDER]
    row = [tanggal_str] + [int(counters.get(name, 0)) for name in REKAP_ORDER] + jam_joined

    if mode == "xlsx":
        try:
            wb = load_workbook(REKAP_FILE)
            ws = wb.active
            ws.append(row)
            if any(v > 0 for v in row[1:1+len(REKAP_ORDER)]):
                try:
                    update_grafik_sheet(wb)  # type: ignore[name-defined]
                except Exception:
                    pass
            wb.save(REKAP_FILE)
        except Exception:
            with open(REKAP_FILE_CSV, "a", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(row)
    else:
        with open(REKAP_FILE_CSV, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)

# ======= Baca 7 hari =======
def read_last_7_days() -> Tuple[List[str], Dict[str, List[int]]]:
    # Mengambil kolom "Down <name>" berdasarkan header agar tahan terhadap tambahan kolom JamDown
    if EXCEL_AVAILABLE and REKAP_FILE_XLSX.exists():
        try:
            wb = load_workbook(REKAP_FILE_XLSX, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                return [], {name: [] for name in REKAP_ORDER}
            headers = list(rows[0])
            idx_map = {name: (headers.index(f"Down {name}") if f"Down {name}" in headers else None) for name in REKAP_ORDER}
            body = rows[1:]
            last = body[-7:]
            dates = [str(r[0]) for r in last]
            data = {name: [] for name in REKAP_ORDER}
            for r in last:
                for name in REKAP_ORDER:
                    i = idx_map[name]
                    val = 0
                    if i is not None and i < len(r) and isinstance(r[i], (int, float)):
                        val = int(r[i])
                    data[name].append(val)
            return dates, data
        except Exception:
            pass
    if REKAP_FILE_CSV.exists():
        with open(REKAP_FILE_CSV, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        if len(rows) <= 1:
            return [], {name: [] for name in REKAP_ORDER}
        headers = rows[0]
        idx_map = {name: (headers.index(f"Down {name}") if f"Down {name}" in headers else None) for name in REKAP_ORDER}
        body = rows[1:]
        last = body[-7:]
        dates = [r[0] for r in last]
        data = {name: [] for name in REKAP_ORDER}
        for r in last:
            for name in REKAP_ORDER:
                i = idx_map[name]
                try:
                    data[name].append(int(r[i])) if i is not None else data[name].append(0)
                except Exception:
                    data[name].append(0)
        return dates, data
    return [], {name: [] for name in REKAP_ORDER}

# ======= POPUP (GUI) =======
def show_popup(root: "tk.Tk", message: str, bg="#222", fg="#fff", ms: int = POPUP_DURATION_MS):
    if not TK_AVAILABLE:
        return
    top = tk.Toplevel(root)
    top.overrideredirect(True)
    top.configure(bg=bg)
    root.update_idletasks()
    x = root.winfo_rootx() + root.winfo_width() - 330
    y = root.winfo_rooty() + root.winfo_height() - 130
    top.geometry(f"310x90+{x}+{y}")
    lbl = tk.Label(top, text=message, bg=bg, fg=fg, font=("Segoe UI", 10), wraplength=285, justify="left")
    lbl.pack(expand=True, fill="both", padx=12, pady=10)
    top.after(ms, top.destroy)

# ======= APP TKINTER =======
if TK_AVAILABLE:
    # ======= Tray Controller (pystray) =======
    class _TrayController:
        def __init__(self, app: "MonitorApp"):
            self.app = app
            self.icon = None
            self._thread = None
        def _create_image(self):
            img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
            d = ImageDraw.Draw(img)
            d.ellipse((8, 8, 56, 56), fill=(34, 197, 94, 255))  # green dot
            d.ellipse((22, 22, 42, 42), fill=(15, 17, 21, 255))  # hole
            return img
        def _run(self):
            menu = pystray.Menu(
                pystray.MenuItem("Restore", self._on_restore),
                pystray.MenuItem("Exit", self._on_exit),
            )
            self.icon = pystray.Icon("PingMonitor", self._create_image(), "Ping Monitor – Realtime", menu)
            self.icon.run()
        def show(self):
            if self.icon is not None:
                return
            self._thread = threading.Thread(target=self._run, daemon=True)
            self._thread.start()
        def hide(self):
            try:
                if self.icon is not None:
                    self.icon.stop()
            finally:
                self.icon = None
        def _on_restore(self, icon, item):
            self.app.after(0, self.app.restore_from_tray)
        def _on_exit(self, icon, item):
            self.app.after(0, self.app.force_exit)

    class MonitorApp(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Realtime Ping Monitor (Tk)")
            self.geometry("600x450")
            self.resizable(False, False)
            self.configure(bg="#0f1115")

            style = ttk.Style(self)
            style.theme_use("clam")
            style.configure("TFrame", background="#0f1115")
            style.configure("TLabel", background="#0f1115", foreground="#e5e7eb")
            style.configure("Header.TLabel", font=("Segoe UI", 11, "bold"))
            style.configure("Muted.TLabel", foreground="#9ca3af")

            wrapper = ttk.Frame(self)
            wrapper.pack(fill="both", expand=True, padx=12, pady=12)

            topbar = ttk.Frame(wrapper)
            topbar.pack(fill="x")
            ttk.Label(topbar, text="Targets", style="Header.TLabel").pack(side="left")
            mode_text = "DUMMY" if USE_DUMMY else "REALTIME"
            self.mode_label = ttk.Label(topbar, text=f"Mode: {mode_text}", style="Muted.TLabel")
            self.mode_label.pack(side="right")
            # Tray controller (Windows + pystray + Pillow)
            self.tray = _TrayController(self) if (TRAY_AVAILABLE and platform.system().lower()=="windows") else None

            hdr = ttk.Frame(wrapper)
            hdr.pack(fill="x", pady=(8, 4))
            ttk.Label(hdr, text=" ", width=2).grid(row=0, column=0, sticky="w")
            ttk.Label(hdr, text="Name", width=18, style="Muted.TLabel").grid(row=0, column=1, sticky="w")
            ttk.Label(hdr, text="Status", width=12, style="Muted.TLabel").grid(row=0, column=2, sticky="w")
            ttk.Label(hdr, text="Down/day", width=10, style="Muted.TLabel").grid(row=0, column=3, sticky="w")

            self.rows: Dict[str, Dict[str, object]] = {}
            for t in TARGETS:
                row = ttk.Frame(wrapper)
                row.pack(fill="x", pady=3)

                dot = tk.Label(row, text="●", fg="#9ca3af", bg="#0f1115", font=("Segoe UI", 18, "bold"))
                dot.grid(row=0, column=0, padx=(0, 12), sticky="w")

                ttk.Label(row, text=t["name"], width=18).grid(row=0, column=1, sticky="w")
                status_lbl = ttk.Label(row, text="...", width=12, font=("Segoe UI", 10, "bold"))
                status_lbl.grid(row=0, column=2, sticky="w")
                cnt_lbl = ttk.Label(row, text="0", width=10, style="Muted.TLabel")
                cnt_lbl.grid(row=0, column=3, sticky="w")

                self.rows[t["name"]] = {
                    "dot": dot,
                    "status": status_lbl,
                    "cnt": cnt_lbl,
                    "host": t["host"],
                    "pulse_job": None,
                    "pulse_big": False,
                }

            ttk.Separator(wrapper).pack(fill="x", pady=8)

            ctr = ttk.Frame(wrapper)
            ctr.pack(fill="x")
            self.footer = ttk.Label(ctr, text="Menunggu update...", style="Muted.TLabel")
            self.footer.pack(side="left")

            btns = ttk.Frame(ctr)
            btns.pack(side="right")
            ttk.Button(btns, text="Chart Mingguan", command=self.open_weekly_chart).pack(side="left", padx=(0, 6))
            ttk.Button(btns, text="Rekap Now", command=self.rekap_now).pack(side="left", padx=(0, 6))
            ttk.Button(btns, text="Open Folder", command=self.open_folder).pack(side="left")

            self.last_status = {t["name"]: "INIT" for t in TARGETS}
            self.incident_open = {t["name"]: False for t in TARGETS}
            self.down_counter = {t["name"]: 0 for t in TARGETS}
            self.down_times: Dict[str, List[str]] = {t["name"]: [] for t in TARGETS}
            self.last_rekap_day: date | None = None

            ensure_workbook(REKAP_FILE)

            self.executor = ThreadPoolExecutor(max_workers=min(8, len(TARGETS)))
            self._tick_lock = threading.Lock()

            self.after(200, self.tick)

            # Keep Awake aktif jika diminta
            if KEEP_AWAKE:
                set_keep_awake(True)

        # ===== Helpers =====
        def set_dot(self, name: str, status: str):
            info = self.rows[name]
            dot: tk.Label = info["dot"]  # type: ignore
            dot.configure(fg=color_for(status))

        def start_pulse(self, name: str):
            info = self.rows[name]
            if info["pulse_job"]:
                return
            def _step():
                big = not info["pulse_big"]
                info["pulse_big"] = big
                dot: tk.Label = info["dot"]  # type: ignore
                size = 22 if big else 18
                dot.configure(font=("Segoe UI", size, "bold"))
                job = self.after(380, _step)
                info["pulse_job"] = job
            _step()

        def stop_pulse(self, name: str):
            info = self.rows[name]
            job = info.get("pulse_job")
            if job:
                try:
                    self.after_cancel(job)  # type: ignore
                except Exception:
                    pass
            info["pulse_job"] = None
            info["pulse_big"] = False
            dot: tk.Label = info["dot"]  # type: ignore
            dot.configure(font=("Segoe UI", 18, "bold"))

        def reset_counter(self):
            for k in self.down_counter:
                self.down_counter[k] = 0
                self.down_times[k] = []
                self.rows[k]["cnt"].configure(text="0")  # type: ignore
            show_popup(self, "Counter harian di-reset", bg="#111827", fg="#e5e7eb")

        def open_folder(self):
            try:
                if platform.system().lower() == "windows":
                    os.startfile(str(REKAP_DIR))  # type: ignore[attr-defined]
                else:
                    messagebox.showinfo("Info", f"Folder: {REKAP_DIR}")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal buka folder:\n{e}")

        def rekap_now(self):
            now = datetime.now()
            write_rekap(now.date(), self.down_counter, self.down_times)
            for k in self.down_counter:
                self.down_counter[k] = 0
                self.down_times[k] = []
                self.rows[k]["cnt"].configure(text="0")  # type: ignore
            self.last_rekap_day = now.date()
            show_popup(self, f"Rekap tersimpan:\n{REKAP_FILE}", bg="#111827", fg="#e5e7eb", ms=2200)

        def schedule_rekap_if_needed(self, now: datetime):
            wd = now.weekday()  # Mon=0..Sun=6
            need = False
            if 0 <= wd <= 4 and now.hour == 15 and now.minute == 30:
                need = True
            elif wd == 5 and now.hour == 13 and now.minute == 30:
                need = True
            if need and self.last_rekap_day != now.date():
                write_rekap(now.date(), self.down_counter, self.down_times)
                for k in self.down_counter:
                    self.down_counter[k] = 0
                    self.down_times[k] = []
                    self.rows[k]["cnt"].configure(text="0")  # type: ignore
                self.last_rekap_day = now.date()
                show_popup(self, f"Rekap tersimpan:\n{REKAP_FILE}", bg="#111827", fg="#e5e7eb", ms=2200)

        def open_weekly_chart(self):
            if not (TK_AVAILABLE and MPL_AVAILABLE):
                messagebox.showinfo("Chart", "Matplotlib/Tkinter tidak tersedia.")
                return
            dates, data = read_last_7_days()
            if not dates:
                messagebox.showinfo("Chart Mingguan", "Belum ada data rekap untuk ditampilkan.")
                return
            win = tk.Toplevel(self)
            win.title("Chart Mingguan - 7 Hari Terakhir")
            win.geometry("860x480")
            win.configure(bg="#0f1115")
            win.resizable(False, False)
            fig = Figure(figsize=(8.6, 4.2), dpi=100)
            ax = fig.add_subplot(111)
            for name in REKAP_ORDER:
                if data.get(name):
                    ax.plot(dates, data[name], marker="o", linewidth=2, label=name)
            ax.set_title("Jumlah Kejadian DOWN per Hari", fontsize=12)
            ax.set_xlabel("Tanggal")
            ax.set_ylabel("Count DOWN")
            ax.grid(True, linestyle="--", alpha=0.35)
            ax.legend(loc="upper left", ncol=2, fontsize=8)
            fig.tight_layout()
            canvas = FigureCanvasTkAgg(fig, master=win)
            canvas_widget = canvas.get_tk_widget()
            canvas_widget.pack(fill="both", expand=True, padx=10, pady=10)
            canvas.draw()
            bar = ttk.Frame(win)
            bar.pack(fill="x", padx=10, pady=(0, 10))
            ttk.Label(bar, text=f"Sumber: {REKAP_FILE}", style="Muted.TLabel").pack(side="left")

        def tick(self):
            if not self._tick_lock.acquire(blocking=False):
                self.after(200, self.tick)
                return
            now = datetime.now()

            results: Dict[str, str] = {}
            if USE_DUMMY:
                for name in self.rows.keys():
                    results[name] = dummy_status()
            else:
                futures = {self.executor.submit(check_target, info["host"]): name for name, info in self.rows.items()}
                for fut in as_completed(futures):
                    name = futures[fut]
                    try:
                        results[name] = fut.result()
                    except Exception:
                        results[name] = "DOWN"

            for name, info in self.rows.items():
                st = results.get(name, "DOWN")
                self.set_dot(name, st)
                info["status"].configure(text=st)  # type: ignore

                before = self.last_status[name]
                self.last_status[name] = st

                if before != "DOWN" and st == "DOWN" and not self.incident_open[name]:
                    self.incident_open[name] = True
                    self.down_counter[name] += 1
                    self.down_times[name].append(now.strftime("%H:%M"))
                    info["cnt"].configure(text=str(self.down_counter[name]))  # type: ignore
                    self.start_pulse(name)
                    show_popup(self, f"ALERT: {name} DOWN", bg="#7f1d1d", fg="#fff")

                if self.incident_open[name] and st != "DOWN":
                    self.incident_open[name] = False
                    self.stop_pulse(name)
                    show_popup(self, f"Pulih: {name} {st}", bg="#064e3b", fg="#d1fae5")

            self.footer.configure(text=f"Last update: {now.strftime('%H:%M:%S')}")
            self.schedule_rekap_if_needed(now)

            self._tick_lock.release()
            self.after(int(INTERVAL_SEC * 1000), self.tick)

        def on_close(self):
            try:
                self.executor.shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass
            # Matikan Keep Awake saat keluar
            try:
                if _KEEP_AWAKE_ON:
                    set_keep_awake(False)
            except Exception:
                pass
            # Tutup tray jika ada
            try:
                if getattr(self, "tray", None):
                    self.tray.hide()
            except Exception:
                pass
            self.destroy()

        def on_close_to_tray(self):
            """Klik X -> ke tray (jika tersedia); jika tidak, minimize ke taskbar."""
            if getattr(self, "tray", None):
                try:
                    self.withdraw()
                except Exception:
                    pass
                self.tray.show()
            else:
                self.iconify()
                show_popup(self, "Berjalan di background (tray tidak tersedia)")

        def restore_from_tray(self):
            try:
                if getattr(self, "tray", None):
                    self.tray.hide()
                self.deiconify()
                self.lift()
                self.focus_force()
            except Exception:
                pass

        def force_exit(self):
            self.on_close()

# ======= HEADLESS (CLI) =======
class HeadlessPingMonitor:
    def __init__(self):
        self.last_status = {t["name"]: "INIT" for t in TARGETS}
        self.incident_open = {t["name"]: False for t in TARGETS}
        self.down_counter = {t["name"]: 0 for t in TARGETS}
        self.down_times: Dict[str, List[str]] = {t["name"]: [] for t in TARGETS}
        self.last_rekap_day: date | None = None
        ensure_workbook(REKAP_FILE)
        self.executor = ThreadPoolExecutor(max_workers=min(8, len(TARGETS)))

    def _get_results(self) -> Dict[str, str]:
        if USE_DUMMY:
            return {t["name"]: dummy_status() for t in TARGETS}
        futures = {self.executor.submit(check_target, t["host"]): t["name"] for t in TARGETS}
        results: Dict[str, str] = {}
        for fut in as_completed(futures):
            name = futures[fut]
            try:
                results[name] = fut.result()
            except Exception:
                results[name] = "DOWN"
        return results

    def schedule_rekap_if_needed(self, now: datetime):
        wd = now.weekday()
        need = False
        if 0 <= wd <= 4 and now.hour == 15 and now.minute == 30:
            need = True
        elif wd == 5 and now.hour == 13 and now.minute == 30:
            need = True
        if need and self.last_rekap_day != now.date():
            write_rekap(now.date(), self.down_counter, self.down_times)
            for k in self.down_counter:
                self.down_counter[k] = 0
                self.down_times[k] = []
            self.last_rekap_day = now.date()
            print(f"[REKAP] Tersimpan -> {REKAP_FILE}")

    def run(self, iterations: int = 10):
        print("[HEADLESS] Tkinter tidak tersedia. Menjalankan mode CLI.")
        print(f"Base dir: {REKAP_DIR}")
        try:
            if KEEP_AWAKE:
                set_keep_awake(True)
            for i in range(iterations):
                now = datetime.now()
                results = self._get_results()
                line = [now.strftime("%H:%M:%S")] + [f"{name}:{results.get(name)}" for name in REKAP_ORDER]
                print(" | ".join(line))

                for name in REKAP_ORDER:
                    st = results.get(name, "DOWN")
                    before = self.last_status[name]
                    self.last_status[name] = st
                    if before != "DOWN" and st == "DOWN" and not self.incident_open[name]:
                        self.incident_open[name] = True
                        self.down_counter[name] += 1
                        self.down_times[name].append(now.strftime("%H:%M"))
                        print(f"  ALERT: {name} DOWN -> count={self.down_counter[name]}")
                    if self.incident_open[name] and st != "DOWN":
                        self.incident_open[name] = False
                        print(f"  Pulih: {name} {st}")

                self.schedule_rekap_if_needed(now)
                time.sleep(INTERVAL_SEC)
        finally:
            # reset Keep Awake
            try:
                if _KEEP_AWAKE_ON:
                    set_keep_awake(False)
            except Exception:
                pass
        print("[HEADLESS] Selesai.")

# ======= TESTS =======
import tempfile

def _set_rekap_base_dir_for_tests(tmp: Path):
    global REKAP_DIR, REKAP_FILE_XLSX, REKAP_FILE_CSV, REKAP_FILE
    REKAP_DIR = tmp
    REKAP_DIR.mkdir(parents=True, exist_ok=True)
    REKAP_FILE_XLSX = REKAP_DIR / "rekap_ping.xlsx"
    REKAP_FILE_CSV = REKAP_DIR / "rekap_ping.csv"
    REKAP_FILE = REKAP_FILE_XLSX if EXCEL_AVAILABLE else REKAP_FILE_CSV


def run_tests():
    print("[TEST] Start")
    # Test color map
    assert color_for("UP").startswith("#"), "color_for UP failed"
    assert color_for("WARNING").startswith("#"), "color_for WARNING failed"
    assert color_for("DOWN").startswith("#"), "color_for DOWN failed"

    # Test ensure_workbook + write_rekap (XLSX atau CSV) + Grafik (jika Excel)
    with tempfile.TemporaryDirectory() as d:
        _set_rekap_base_dir_for_tests(Path(d))
        ensure_workbook(REKAP_FILE)
        assert REKAP_FILE.exists(), "Rekap file tidak dibuat"
        # Tulis satu baris tanpa DOWN -> tidak perlu grafik
        write_rekap(date.today(), {name: 0 for name in REKAP_ORDER}, {name: [] for name in REKAP_ORDER})
        if EXCEL_AVAILABLE:
            wb = load_workbook(REKAP_FILE)
            assert "Grafik" not in wb.sheetnames or wb["Grafik"].max_row <= 1
        # Tulis baris dengan DOWN -> Grafik wajib ada (Excel)
        write_rekap(date.today(), {name: (1 if i % 2 == 0 else 0) for i, name in enumerate(REKAP_ORDER)}, {name: ["09:30"] for name in REKAP_ORDER})
        if EXCEL_AVAILABLE:
            wb = load_workbook(REKAP_FILE)
            assert "Grafik" in wb.sheetnames, "Sheet Grafik tidak dibuat"

    # Distribusi dummy minimal
    outs = {dummy_status() for _ in range(200)}
    assert {"UP", "WARNING", "DOWN"}.intersection(outs), "dummy_status tidak variatif"

    # Pastikan flag CLOSE_TO_TRAY selalu tersedia
    assert "CLOSE_TO_TRAY" in globals(), "CLOSE_TO_TRAY missing"

    print("[TEST] OK")

# ======= Entrypoint =======
if __name__ == "__main__":
    if os.environ.get("RUN_TESTS") == "1":
        run_tests()
        sys.exit(0)

    if TK_AVAILABLE:
        app = MonitorApp()  # type: ignore[name-defined]
        handler = app.on_close_to_tray if globals().get("CLOSE_TO_TRAY", True) else app.on_close
        app.protocol("WM_DELETE_WINDOW", handler)
        app.mainloop()
    else:
        HeadlessPingMonitor().run(iterations=int(os.environ.get("ITERATIONS", "10")))
